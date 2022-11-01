import os
import openpyxl
from openpyxl.comments import Comment
from openpyxl.chart import ScatterChart, Reference, Series  # グラフの形式、系列の扱い
from openpyxl.chart.axis import DateAxis

filepath = 'C:/Users/'+os.getlogin()+'/Downloads/LUPOUT/'
allfilename_list = os.listdir(filepath)  # ディレクトリ内のファイル名を全て取得
LUPOUTt_xlsxfilename_list = []
filename_list = []   # LUPOUTtのみを格納
for fname in allfilename_list:
    if "_LUPOUTt.log" in fname:
        filename_list.append(fname)
print(filename_list)

m_wb = openpyxl.Workbook()  # 最終的にマージする先のエクセルブックを作成
m_ws = m_wb.active
m_wb.remove(m_ws) # ワークブックを作ったときに最初からあるワークシートを削除

for fname in filename_list:
    target_file = filepath + fname
    tmp_ene_list = [];  tmp_spin_list = []
    with open(target_file, encoding='utf-8') as f:
        tmp_atomnum = 0;  tmp_nodenum = 0;  atomcount = 0
        while line := f.readline():  # 1行ずつ読み込み (セイウチ演算子)
            if "#" in line:
                tmp_nodenum += 1
                while 1 == 1:
                    line = f.readline()
                    elem = line.split()
                    if "Item" in line or "=" in line or "ENE" in line or len(elem) < 4:
                        break
                    elif atomcount == 0:
                        tmp_atomnum += 1
                atomcount = 1
            elem = line.split()
            if "ENE" in line or "Ene" in line:
                tmp_ene_list.append(elem[1])
            if "SPIN" in line or "Spin" in line:
                tmp_spin_list.append(elem[1])

    # xlsxファイル作成、エネルギー書き込み
    jobname = fname.replace("_LUPOUTt.log","")
    target_xlsxfile = filepath + fname.replace(".log", ".xlsx")
    LUPOUTt_xlsxfilename_list.append(target_xlsxfile)
    wb = openpyxl.Workbook()   # ブックオブジェクトを新規作成
    wb.create_sheet(jobname)  # ブック内に jobname という名前のシートを新規作成
    ws = wb[jobname]  # ブック内の jobname という名前のシートをオブジェクトとして開く
    ws["A1"] = fname  # A1要素にファイル名を出力
    ws[ws.cell(column=1,row=2).coordinate] = "NODE"  # 1列目2行目に "NODE" を出力
    for i in range(tmp_nodenum):
        ws[ws.cell(column=1,row=3+i).coordinate] = int(i)  # 1列目3+i行目に # NODE を出力
    ws[ws.cell(column=2,row=2).coordinate] = "ENERGY"  # 2列目2行目に "ENERGY" を出力
    for i in range(tmp_nodenum):
        ws[ws.cell(column=2,row=3+i).coordinate] = float(tmp_ene_list[i])  # 2列目3+i行目に絶対値のエネルギー(hartree)を出力
    ws[ws.cell(column=3,row=2).coordinate] = "ΔE"  # 3列目2行目に "ΔE" を出力
    for i in range(tmp_nodenum):
        ws[ws.cell(column=3,row=3+i).coordinate] = 2625.5*(float(tmp_ene_list[i])-float(tmp_ene_list[0]))  # 3列目3+i行目に相対エネルギー(kJ/mol)を出力
    ws[ws.cell(column=4,row=2).coordinate] = "Spin(**2)"  # 4列目2行目に "Spin(**2)" を出力
    for i in range(tmp_nodenum):
        ws[ws.cell(column=4,row=3+i).coordinate] = tmp_spin_list[i]  # 4列目3+i行目にスピン二乗値を出力

    # グラフオブジェクト作成
    graph_obj = openpyxl.chart.LineChart()  # 折れ線グラフとしてグラフオブジェクトを生成
    graph_obj.title  = fname                # グラフのタイトル
    graph_obj.x_axis.title = 'NODE'         # x軸のタイトル
    graph_obj.y_axis.title = 'Energy'       # y軸のタイトル
    graph_obj.style  = 12                   # グラフのフォントサイズ
    graph_obj.height = 10                   # 高さ
    graph_obj.width  = 15                   # 幅

    # データの範囲設定
    y_values = Reference(ws, min_col=3, min_row=3, max_col=3, max_row=3+tmp_nodenum)  # y軸範囲 (3列目の相対エネルギー(kJ/mol)を出力)
    graph_obj.add_data(y_values, titles_from_data=True)
    x_values = Reference(ws, min_col=1, min_row=3, max_col=1, max_row=3+tmp_nodenum)  # x軸範囲 (1列目の # NODE を出力)
    graph_obj.set_categories(x_values)

    # グラフのスタイル変更・配置
    style = graph_obj.series[0]
    style.graphicalProperties.line.solidFill = "6495ed"   # 折れ線の色
    style.graphicalProperties.line.width = 20000          # グラフの幅
    style.smooth = False                                  # スムージング処理
    ws.add_chart(graph_obj, "F3")                         # ブック中のF3セルの位置にグラフを配置
    wb.remove(wb['Sheet'])                                # ブック中の"Sheet"という名前のシートを削除

    # ワークブックオブジェクトにデータを書き出す
    # wb.save(target_xlsxfile)  # エクセルファイルとして保存（オプション）
    m_wb._sheets.append(wb._sheets[0])  # マージ先のワークブックオブジェクトに格納

m_wb.save(filepath+'LUPOUTt_all.xlsx')  # すべてのLUPOUTt.logのデータをマージしたエクセルファイルを保存

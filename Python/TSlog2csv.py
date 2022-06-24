import sys
import openpyxl
from openpyxl.comments import Comment

# args = sys.argv

filepath = 'C:/Users/xxxx/Downloads/saddle/'
filename_list = ["test_DS1.log"] # sys.argv
# filepath = ''
# filename_list = args # sys.argv
column_title_list = ["FW", "TS", "BW"]

# open a new excel file
target_xlsxfile = filepath + "all.xlsx"
wb = openpyxl.Workbook()
for i in range(len(filename_list)):
    target_file = filepath + filename_list[i]
    TS_structure_list = []; FW_structure_list = []; BW_structure_list = []; AppEQ_structure_list = []
    TS_therm_list = []; FW_therm_list = []; BW_therm_list = []; AppEQ_therm_list = []
    TS_replaced_therm_list = []; FW_replaced_therm_list = []; BW_replaced_therm_list = []; AppEQ_replaced_therm_list = []
    with open(target_file, encoding='utf-8') as f:
        SADDLE_mode = 0
        LUP_mode = 0
        DS_mode = 0
        atomnum = 0
        ts_freq = 0
        ts_mode = 1     # 1: TS
        irc_mode = 0    # 1: FW,  2: BW
        eigencheck = 0  # if eigencheck is valid (1) or not (0)
        AppEQ_mode = 0
        while line := f.readline():  # read lines one by one
            # JOBtype check
            if "(by SM)" in line:
                if "SADDLE:" in line:
                    SADDLE_mode = 1
                elif "LUP:" in line:
                    LUP_mode = 1
                elif "DS-AFIR:" in line:
                    DS_mode = 1

            # atomnum counting
            if '# ' in line and atomnum == 0:
                while 1 == 1:
                    line = f.readline()
                    elem = line.split()
                    if "Item" in line or "=" in line or "ENE" in line or len(elem) < 4:
                        break
                    else:
                        atomnum += 1
                # print("atomnum counting done.")

            if "Maximum number of iteration was exceeded" in line:
                if ts_mode == 1:     # TS
                    TS_structure_list.append([])
                    FW_structure_list.append([])
                    BW_structure_list.append([])
                    TS_therm_list.append([])
                    FW_therm_list.append([])
                    BW_therm_list.append([])
                    TS_replaced_therm_list.append([])
                    FW_replaced_therm_list.append([])
                    BW_replaced_therm_list.append([])
                elif irc_mode == 1:    # FW
                    FW_structure_list.append([])
                    BW_structure_list.append([])
                    FW_therm_list.append([])
                    BW_therm_list.append([])
                    FW_replaced_therm_list.append([])
                    BW_replaced_therm_list.append([])
                elif irc_mode == 2:    # BW
                    BW_structure_list.append([])
                    BW_therm_list.append([])
                    BW_replaced_therm_list.append([])
                    irc_mode = 0
                # print("Max # of ITR...")

            if "Start MIN-optimization of AppEQ" in line:
                AppEQ_mode = 1

            if "Optimized structure" in line:
                # print("Optimized")
                tmp_structure_list = []
                for j in range(atomnum):
                    line = f.readline()
                    elem = line.split()
                    tmp_structure_list.append(line[:-1].replace("\t", ""))
                if AppEQ_mode == 0:
                    if ts_mode == 1:     # TS
                        TS_structure_list.append(tmp_structure_list)
                        # print("Opt TSstr found")
                    elif irc_mode == 1:    # FW
                        FW_structure_list.append(tmp_structure_list)
                        # print("Opt FWstr found")
                    elif irc_mode == 2:    # BW
                        BW_structure_list.append(tmp_structure_list)
                        # print("Opt BWstr found")
                else:
                    AppEQ_structure_list.append(tmp_structure_list)
                    # print("Opt AppEQstr found")

            if "(FORWARD)" in line:
                # print("FORWARD")
                irc_mode = 1
                ts_mode = 0
                ts_freq = 0

            if "(BACKWARD)" in line:
                # print("BACKWARD")
                irc_mode = 2

            if "Thermochemistry" in line:
                eigencheck = 1
                tmp_therm_list = []
                if "(use all positive eigenvalues)" in line:
                    # print("all")
                    for j in range(14):
                        line = f.readline()
                        elem = line.split()
                        if j == 13:
                            tmp_therm_list.append(float(elem[3]))
                        else:
                            tmp_therm_list.append(float(elem[2]))
                    if AppEQ_mode == 0:
                        if ts_mode == 1 and ts_freq < 2:    # TS
                            TS_therm_list.append(tmp_therm_list)
                            ts_freq += 1
                        elif irc_mode == 1:    # FW
                            FW_therm_list.append(tmp_therm_list)
                        elif irc_mode == 2:    # BW
                            BW_therm_list.append(tmp_therm_list)
                    else:
                        AppEQ_therm_list.append(tmp_therm_list)

                elif "(after the above replacements)" in line:
                    # print("after")
                    for j in range(14):
                        line = f.readline()
                        elem = line.split()
                        if j == 13:
                            tmp_therm_list.append(float(elem[3]))
                        else:
                            tmp_therm_list.append(float(elem[2]))
                    if AppEQ_mode == 0:
                        if ts_mode == 1 and ts_freq < 2:    # TS
                            TS_replaced_therm_list.append(tmp_therm_list)
                            ts_freq += 1
                        elif irc_mode == 1:    # FW
                            FW_replaced_therm_list.append(tmp_therm_list)
                        elif irc_mode == 2:    # BW
                            BW_replaced_therm_list.append(tmp_therm_list)
                    else:
                        AppEQ_replaced_therm_list.append(tmp_therm_list)

            if "IRC following along both forward and backward directions were finished" in line:
                ts_mode = 1

            if not line:
                break

    # シートを追加
    wb.create_sheet(index=i, title=filename_list[i].replace(".log", ""))
    ws = wb[filename_list[i].replace(".log", "")]
    ws["A1"] = filename_list[i]

    # セル＆コメントの書き込み
    TS_num = len(TS_structure_list)
    for j in range(TS_num):
        ws[ws.cell(row=2,column=1+4*j).coordinate] = "AppTS " + str(j)
        if len(TS_structure_list[j]) != 0:
            ws[ws.cell(row=3,column=1+4*j).coordinate] = "FW"
            tmp_comment_str = FW_structure_list[j][0]
            for k in range(len(FW_structure_list[j])-1):
                tmp_comment_str = tmp_comment_str + '\n' + FW_structure_list[j][k+1]
            comment = Comment(tmp_comment_str, 'Comment Author')
            comment.width = 400
            comment.height = 400
            ws[ws.cell(row=3,column=1+4*j).coordinate].comment = comment
            if eigencheck == 1:
                for k in range(len(FW_therm_list[j])):
                    ws[ws.cell(row=4+k,column=1+4*j).coordinate] = FW_therm_list[j][k]
                for k in range(len(FW_replaced_therm_list[j])):
                    ws[ws.cell(row=21+k,column=1+4*j).coordinate] = FW_replaced_therm_list[j][k]

            ws[ws.cell(row=3,column=2+4*j).coordinate] = "TS"
            tmp_comment_str = TS_structure_list[j][0]
            for k in range(len(TS_structure_list[j])-1):
                tmp_comment_str = tmp_comment_str + '\n' + TS_structure_list[j][k+1]
            comment = Comment(tmp_comment_str, 'Comment Author')
            comment.width = 400
            comment.height = 400
            ws[ws.cell(row=3,column=2+4*j).coordinate].comment = comment
            if eigencheck == 1:
                for k in range(len(TS_therm_list[j])):
                    ws[ws.cell(row=4+k,column=2+4*j).coordinate] = TS_therm_list[j][k]
                for k in range(len(TS_replaced_therm_list[j])):
                    ws[ws.cell(row=21+k,column=2+4*j).coordinate] = TS_replaced_therm_list[j][k]

            ws[ws.cell(row=3,column=3+4*j).coordinate] = "BW"
            tmp_comment_str = BW_structure_list[j][0]
            for k in range(len(BW_structure_list[j])-1):
                tmp_comment_str = tmp_comment_str + '\n' + BW_structure_list[j][k+1]
            comment = Comment(tmp_comment_str, 'Comment Author')
            comment.width = 400
            comment.height = 400
            ws[ws.cell(row=3,column=3+4*j).coordinate].comment = comment
            if eigencheck == 1:
                for k in range(len(BW_therm_list[j])):
                    ws[ws.cell(row=4+k,column=3+4*j).coordinate] = BW_therm_list[j][k]
                for k in range(len(BW_replaced_therm_list[j])):
                    ws[ws.cell(row=21+k,column=3+4*j).coordinate] = BW_replaced_therm_list[j][k]

    AppEQ_num = len(AppEQ_structure_list)
    for j in range(AppEQ_num):
        ws[ws.cell(row=3,column=4*(TS_num)+1+j).coordinate] = "AppEQ " + str(j)
        tmp_comment_str = AppEQ_structure_list[j][0]
        for k in range(len(AppEQ_structure_list[j])-1):
            tmp_comment_str = tmp_comment_str + '\n' + AppEQ_structure_list[j][k+1]
        comment = Comment(tmp_comment_str, 'Comment Author')
        comment.width = 400
        comment.height = 400
        ws[ws.cell(row=3,column=4*(TS_num)+1+j).coordinate].comment = comment
        if eigencheck == 1:
            for k in range(len(AppEQ_therm_list[j])):
                ws[ws.cell(row=4+k,column=4*(TS_num)+1+j).coordinate] = AppEQ_therm_list[j][k]
            for k in range(len(AppEQ_replaced_therm_list[j])):
                ws[ws.cell(row=21+k,column=4*(TS_num)+1+j).coordinate] = AppEQ_replaced_therm_list[j][k]

    print("done:", target_file)

# save the excel file
wb.save(target_xlsxfile)
print("csv->", target_xlsxfile)
